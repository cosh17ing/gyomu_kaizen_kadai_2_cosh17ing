# %%
#ライブラリのインポート
import TkEasyGUI as sg  #
import os, sys
import calendar, datetime
import jpbizday  #
import chardet  #
import openpyxl.styles
import pandas as pd     #
import re
import math
import openpyxl #

# %%
#ファイルを読み込み、読み込むファイルのエンコードを判別する
#改めてファイルを読み込み、データフレームにする
#それぞれのデータフレームの氏名欄に\u3000(全角スペース)がある場合、半角スペースに変換する

#この3つを同時に行い、データフィールドを戻り値とする関数
def csvConvert(file_path) :
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    encoding = result['encoding']
    
    df = pd.read_csv(file_path, encoding=encoding)

    if '\u3000' in df.columns.values[0]:
        df = df.rename(columns={df.columns.values[0] : df.columns.values[0].replace('\u3000',' ')})

    return df

# %%
# strとfloat_nanを同時に判別する関数
# nanならFalse、nanでないならその値を返す
def nan_check(val) :
    if type(val) is str :return val
    else :
        if math.isnan(val) :
            return False
        else :return val

# masterデータが存在すればそれを返し、存在しなければ新たに作成し、保存する
def checkMasterCSV(file_path) :

    # ファイルの存在確認
    if os.path.exists(file_path):
        # ファイルが存在すれば読み込む
        #！！！readerror(?)とpermissionerrorでtryを通すこと！！！
        masterdf = pd.read_csv(file_path, index_col=0, dtype={'受給者証番号':'str', '月間アセスメント担当者名':'str'})
        
    else:
        # ファイルが存在しなければ新しいデータフレームを作成
        columns = ['受給者証番号', '月間アセスメント担当者名']
        masterdf = pd.DataFrame(columns=columns)
        
        # フォルダが無ければ作成
        # 相対パスにファイル名のみを指定した場合、フォルダ名は''となる
        if os.path.dirname(file_path) != '' :
            if os.path.exists(os.path.dirname(file_path)) == False :
                os.makedirs(os.path.dirname(file_path))
        # データフレームをcsvファイルとして保存
        masterdf.to_csv(file_path, index=True)

    return masterdf

# %%
# マスターデータ照合
def findMaster(masterdf, searchword) :

    # indexをキーとしてデータフィールド内を検索
    if searchword in masterdf.index.values :
        # 存在する場合、その行の他の列の要素を取得
        result_row = masterdf[masterdf.index == searchword]
        id = str(result_row.values[0][0])
        assessmentAuthor = str(result_row.values[0][1])
        statusUpdate('氏名より受給者証番号および月間アセスメント担当者名を読み込みました。\n\
変更する場合は、受給者証番号または月間アセスメント担当者名の欄を変更して、「実行」ボタンを押してください。', 1)
        return True, id, assessmentAuthor

    # 存在しない場合、返り値はFalse, '', ''
    statusUpdate('その氏名は登録されていません。\n\
受給者証番号および月間アセスメント担当者名を入力してください。', 1)
    return False, '', ''

# %%
# マスターデータ編集
def editMaster(masterdf, name, id, assessmentAuthor, masterDataPath) :
    # masterデータを編集
    masterdf.loc[name] = [id, assessmentAuthor]

    # masterデータを保存
    if os.path.dirname(masterDataPath) != '' :
        if os.path.exists(os.path.dirname(masterDataPath)) == False :
            os.makedirs(os.path.dirname(masterDataPath))
    # データフレームをcsvファイルとして保存
    masterdf.to_csv(masterDataPath, index=True)

    # 返り値はmasterデータ
    return masterdf

# %%
# セルの位置リスト作成
def listCellPosition(year, month, popCheckStatus) :
    # nには曜日に対応したint型の整数が入る　0→月曜日　6→日曜日
    dt = datetime.datetime(year, month, 1)
    n = dt.weekday()
    cell_list_startplace = n

    # nが5以上になると7を引き、-2から再びカウントが進む
    if n > 4 :
        n -= 7
        cell_list_startplace = 0

    # 1日が土曜日で開所日の場合、記入位置を1週分（5箇所）後ろにずらす
    if popCheckStatus : cell_list_startplace += 5

    # cell_list_startplaceがここで決定

    # 月末日を求める
    lastdate = datetime.date(year, month, calendar.monthrange(year, month)[1]).day

    day_list = []

    # nが5以上になると7を引き、-2から再びカウントが進む
    # nが0以上の時にリストに追加する
    for i in range(1, lastdate + 1, 1) :
        if n >= 0 :
            day_list.append(i)

        n += 1
        if n > 4 : n -= 7

    # day_listがここで完成
    
    column_list = ['B', 'C', 'E', 'F', 'G', 'I', 'O', 'P', 'R', 'S', 'T', 'V']
    # row_listは必ずintで扱うこと！！
    row_list = list(range(11, 21, 2))*2 + list(range(28, 38, 2))*2 + list(range(45, 55, 2))*2
    cell_list = []

    for i, r in enumerate(row_list) :
        # B, B, B, B, B, O, O, O, O, O, B, B, ...
        column = 'B'
        if int(i / 5) % 2 == 1 :
            column = 'O'
        
        # 列名と行番号を結合して二次元リストに追加
        cell_list.append([column_list[column_list.index(column)  ] + str(r),
                        column_list[column_list.index(column) + 1] + str(r),
                        column_list[column_list.index(column) + 2] + str(r),
                        column_list[column_list.index(column) + 2] + str(r + 1),
                        column_list[column_list.index(column) + 3] + str(r),
                        column_list[column_list.index(column) + 3] + str(r + 1),
                        column_list[column_list.index(column) + 4] + str(r),
                        column_list[column_list.index(column) + 5] + str(r)])
    
    return day_list, cell_list_startplace, cell_list

# %%
#メイン処理関数
def convertToXls(caseDailyPath, userCaseDailyPath, formatPath, assessmentDay, id,
                 assessmentAuthor, popCheckStatus) :
    # %%
    #ファイルを読み込み、エンコードをUTF-8に変換してデータフレームにする
    #それぞれのデータフレームの氏名欄に\u3000(全角スペース)がある場合、半角スペースに変換する
    caseDaily = csvConvert(caseDailyPath)
    userCaseDaily = csvConvert(userCaseDailyPath)

    name = caseDaily.columns.values[0]
    df = pd.merge(caseDaily, userCaseDaily, on=name, how='left')

    #Excelブックを開く
    try :
        wb = openpyxl.load_workbook(formatPath)
    except PermissionError :
        statusUpdate('フォーマットとなるExcelブックを開くことが出来ませんでした。\n\
Excelブックを閉じてからやり直してください。', 2)
        return 1

    # 年月日の区切り文字を特定する
    separate_code = df[name][0][4]

    # 年月日を取得する
    yyyymm = df[name][0]
    yyyymm_list = yyyymm.split(separate_code)
    yyyymm_list[1] = str(int(yyyymm_list[1]))
    yyyymm_list[2] = str(int(yyyymm_list[2]))
    date = f'　{yyyymm_list[0][-2:]}　年　{yyyymm_list[1]}　月'

    # シートの上部分を記入する
    sheet = wb['在宅支援利用記録表']
    sheet['A1'] = date
    sheet['D2'] = name
    sheet['D3'] = assessmentDay
    sheet['G2'] = id
    sheet['G3'] = assessmentAuthor

    # セルの位置リスト作成
    year = int(yyyymm_list[0])
    month = int(yyyymm_list[1])

    day_list, cell_list_startplace, cell_list = listCellPosition(
                                    year=year, month=month, popCheckStatus=popCheckStatus)
    
    ## cell_list[入力箇所0~29][日ごとの入力欄0~7]

    # day_listでループ　営業日なら日付と欠席を入れる
    for i, d in enumerate(day_list) :
        if jpbizday.is_bizday(datetime.date(year, month, d)) :
            # 日付取得
            date = f'{month}/{d}'
            
            # セル位置設定
            day_position = day_list.index(d)
            cell_position = cell_list[day_position + cell_list_startplace][:]

            sheet[cell_position[0]] = date
            
            sheet[cell_position[1]] = '　　　　　　　　　欠席'
    

    # データを取り出す
    for i1, i2, i3, i4, i5, i6, i7, i8, i9 in zip(
        df[name], df['実績記録票備考欄'], df['出欠等'], df['実績開始時間_x'], df['実績終了時間_x'], 
        df['体温'], df['日報_x'], df['午前のプログラム詳細'], df['午後1のプログラム詳細']
    ):
        yyyymm_list = i1.split(separate_code)
        day = int(yyyymm_list[2])

        # データの行が土曜日の場合、処理を省略する
        if not jpbizday.is_bizday(datetime.date(year, month, day)) :
            continue

        # セル位置設定
        day_position = day_list.index(day)
        cell_position = cell_list[day_position + cell_list_startplace][:]

        # データ処理
        
        # 支援方法
        method = '通所　　　　　　　　　 '
        if i3 == '欠席時対応' :
            method = '　　　　　　　　　欠席'
            continue
        elif i2 == '在宅支援（届出済）' :
            method = '在宅'
        
        # 実績開始時間、実績終了時間
        startTime = i4[:5]
        endTime = i5[:5]
        
        # 体温
        temperature = ''
        if nan_check(i6) :
            # nanじゃないとき
            temperature = f'検温　{str(i6)}　℃'
        #else :
        #    temperature = '検温　　　℃'

        # 体調
        condition = '　　　　普通'
        if '不調' in i7 :
            condition = '　　　　　　　　不良'
        elif '好調' in i7 :
            condition = '良好'

        # 作業活動・訓練等のメニュー
        work = ''
        if not nan_check(i9) :
            # nanのとき
            if nan_check(i8) : work = i8
        else :
            # nanじゃないとき
            work = f'{i8}\n{i9}'
        
        # 支援者コメント
        comment = ''
        result = re.findall(r'「.+?」', i7)
        #print(result)
        if result != [] :
            comment = result[-1][1:-1]
            #print(comment)
        
        # Excelシートに転記
        sheet[cell_position[1]] = method
        sheet[cell_position[2]] = startTime
        sheet[cell_position[3]] = endTime
        if nan_check(i6) : sheet[cell_position[4]] = temperature
        sheet[cell_position[5]] = condition
        # セル内改行を有効にして、水平中央揃え、垂直中央揃えにする
        sheet[cell_position[6]].alignment = openpyxl.styles.Alignment(
            wrapText=True, horizontal='center', vertical='center'
            )
        sheet[cell_position[6]] = work
        sheet[cell_position[7]] = comment

    #Excelブックを保存する
    if not os.path.isdir(val4) :
        os.makedirs(val4)
    try :
        wb.save(savePath)
    except PermissionError :
        statusUpdate('出力先のExcelブックを開くことが出来ませんでした。\n\
Excelブックを閉じてからやり直してください。', 2)
        return 1

    statusUpdate(f'保存しました。\n{savePath}', 2)
    return 0

# 今日の日付を文字列で取得する
def setToday() :
    dt_now = datetime.datetime.now()
    strToday = f'{dt_now.year}年{dt_now.month}月{dt_now.day}日'
    return strToday

strToday = setToday()

# %%
# GUIを定義する
layout = [
    [sg.Text('上4つのテキストボックスを入力し、「抽出」ボタンを押してください。')],
    [sg.Text('caseDaily:'), sg.Input('', key='-caseDailyPath-'),
                            sg.FileBrowse(file_types=(('CSV ファイル', '*.csv'),))],
    [sg.Text('userCaseDaily:'), sg.Input('', key='-userCaseDailyPath-'), 
                                sg.FileBrowse(file_types=(('CSV ファイル', '*.csv'),))],
    [sg.Text('月間アセスメント実施日:'), sg.Input(strToday, key='-assessmentDay-')],
    [sg.Text('Excelブックの出力先:'), sg.Input('', key='-outputPath-'), sg.FolderBrowse()],
    [sg.Button('抽出', key='-searchBtn-')],
    [sg.Text('氏名:'), sg.Input('', key='-name-')],
    [sg.Text('受給者証番号:'), sg.Input('', key='-id-')],
    [sg.Text('月間アセスメント担当者名:'), sg.Input('', key='-assessmentAuthor-')],
    [sg.Checkbox('1日は土曜日で、開所日です。（2週目から記録します。）', key='-popFillPositionCheck-')],
    [sg.Text('', key='-notification1-')],
    [sg.Text('')],
    [sg.Text('Format'), sg.Input('dist/在宅支援記録表及びアセスメント表（週間・月間）.xlsx', 
                                 key='-formatPath-', readonly=True),
                                 sg.FileBrowse(file_types=(('Excel ブック', '*.xlsx'),))],
    [sg.Text('Data'), sg.Input('dist/usermaster.csv', 
                               key='-masterDataPath-', readonly=True),
                               sg.FileSaveAs(file_types=(('CSV ファイル', '*.csv'),))],
    [sg.Button('実行', key='-submitBtn-'), sg.Button('クリア', key='-clearBtn-')],
    [sg.Text('', key='-notification2-')]
]
window = sg.Window('在宅支援記録及びアセスメント表作成支援ツール', layout)

def statusUpdate(text, n) :
    if n == 1 :
        window['-notification1-'].update(text)
    elif n == 2 :
        window['-notification2-'].update(text)
    else :
        window['-notification2-'].update(text + '\n※statusUpdate(**,n)でnの指定が誤っています。1か2を指定してください。')
    return

# %%
#Excelファイル生成
def submit() :
    caseDailyPath = val1
    userCaseDailyPath = val2
    assessmentDay = val3
    name = val5
    id = val6
    assessmentAuthor = val7
    formatPath = valA
    
    # データ参照の処理をここに挟む
    # 再度masterdfデータフィールドを定義する
    masterdf = checkMasterCSV(valB)
    masterdf = editMaster(masterdf=masterdf,
                        name=name,
                        id=val6,
                        assessmentAuthor=val7,
                        masterDataPath=valB)

    convertToXls(caseDailyPath=caseDailyPath,
                userCaseDailyPath=userCaseDailyPath,
                formatPath=formatPath,
                assessmentDay=assessmentDay,
                id=id,
                assessmentAuthor=assessmentAuthor,
                popCheckStatus=popCheckStatus)# :
    #    return
    
    #window['-caseDailyPath-'].update('')
    #window['-userCaseDailyPath-'].update('')

    data_mode = 0
    data_flag = False

    return

# flagは、×ボタンを押した時にFalse、それ以外のボタンを押した時にTrueになる
# GUIの実行において非常に重要なフラグ変数
flag = False
# verify_flagは、テキスト欄の入力が正しいことが確認されればTrueになる
# もっとも、入力に問題があればcontinueで処理を抜けているので、
# どちらかというとセーフガードとして機能している
verify_flag = False
# buffer_flagは、上書き確認・データ変更確認が入るとTrueになり、
# buffer_flagがTrueの状態でテキスト欄が一切変更されずに「実行」ボタンを押すと
# ファイル上書き・マスターデータ変更が実行される
buffer_flag = False
# data_modeは、
# 0:「抽出」ボタンを押していない　「実行」ボタンは作動せず、エラーメッセージを出す
# 1:「抽出」ボタンは押したもののマスターデータ内に対象のデータが存在しないため、新規にデータを追加する
# 2:「抽出」ボタンを押して、マスターデータからデータを参照したことを確認するとTrueになる
# 参照したマスターデータが変更された場合は、処理を中断するがdata_flagをTrueにする
data_mode = 0
# data_flagがTrueのときは、マスターデータの内容の変更を行う
data_flag = False
# bufferは、上書き確認・データ変更確認が入ると、テキスト欄の入力が全てこの変数にリストとして格納される
# テキスト欄が変更されていないことを確認するためのリスト
buffer = []
# databufferは、マスターデータの読み込みに成功すると、その内容がこの変数にリストとして格納される
# マスターデータ関連の入力欄が変更されていないことを確認するためのリスト
dataBuffer = []
# イベントを処理
while True :
    # イベントを取得
    event, values = window.read()

    # ウィンドウが閉じられた時の処理
    if event == sg.WINDOW_CLOSED :
        flag = False
        window.close()
        break

    # 抽出ボタンが押された時の処理
    if event == '-searchBtn-' :
        flag = True

        val1 = values['-caseDailyPath-']
        val2 = values['-userCaseDailyPath-']
        valB = values['-masterDataPath-']

        verify_flag = True

        #Group A
        #??. caseDailyが空欄でない
        #??. userCaseDailyが空欄でない
        #今回は含まない［??. 出力先が空欄でない］
        if val1 == '' or val2 == '' :
            verify_flag = False

            empty_list = []
            if val1 == '' : empty_list.append('caseDaily欄')
            if val2 == '' : empty_list.append('userCaseDaily欄')

            empty_str = ''
            for i, j in enumerate(empty_list) :
                if i > 0 : 
                    #if i == len(empty_list) - 1 :empty_str += 'および' 
                    #else : 
                    empty_str += '、'
                empty_str += j
            statusUpdate(f'{empty_str}は必須項目です。空欄にはできません。', 2)

        if not verify_flag :continue
    
        #Group B
        #??. caseDailyにcaseDailyのファイルが入っている
        #??. caseDailyにcsvファイルが入っている
        #??. userCaseDailyにuserCaseDailyのファイルが入っている
        #??. userCaseDailyにcsvファイルが入っている
        filename1 = val1.split('/').pop()
        filename2 = val2.split('/').pop()

        errMsg_list = []
        if len(filename1) < 4 :
            verify_flag = False
            errMsg_list.append('caseDaily欄に指定した値が有効ではありません。')
        elif filename1[-4:] != '.csv' :
            verify_flag = False
            errMsg_list.append('caseDaily欄に指定したファイルはCSVではありません。')
        elif len(filename1) < 19 :
            verify_flag = False
            errMsg_list.append('caseDaily欄に指定した値が有効ではありません。')
        elif filename1[:9] != 'caseDaily':
            verify_flag = False
            errMsg_list.append('caseDaily欄に指定した値が有効ではありません。')
            
        if len(filename2) < 4 :
            verify_flag = False
            errMsg_list.append('userCaseDaily欄に指定した値が有効ではありません。')
        elif filename2[-4:] != '.csv' :
            verify_flag = False
            errMsg_list.append('userCaseDaily欄に指定したファイルはCSVではありません。')
        elif len(filename2) < 23 :
            verify_flag = False
            errMsg_list.append('userCaseDaily欄に指定した値が有効ではありません。')
        elif filename2[:13] != 'userCaseDaily' :
            verify_flag = False
            errMsg_list.append('userCaseDaily欄に指定した値が有効ではありません。')

        if not verify_flag :
            errMsg = ''
            for i, j in enumerate(errMsg_list) :
                if i : errMsg += '\n'
                errMsg += j
            statusUpdate(errMsg, 2)
            continue
        
        #Group C
        #??. caseDailyとuserCaseDailyのIDが同じである
        #??. caseDailyとuserCaseDailyの年月が同じである
        filename1_list = filename1.split('_')
        filename2_list = filename2.split('_')

        differ_list = []
        if filename1_list[1] != filename2_list[1] or filename1_list[2] != filename2_list[2] :
            verify_flag = False

            if filename1_list[1] != filename2_list[1] : differ_list.append('ID')
            if filename1_list[2][:6] != filename2_list[2][:6] : differ_list.append('年月')

            differ_str = ''
            for i, j in enumerate(differ_list) :
                if i :differ_str += 'および'
                differ_str += j

        if not verify_flag :
            if differ_str == '' :differ_str = 'ファイル名の形式'
            statusUpdate(f'caseDailyとuserCaseDailyで{differ_str}が異なります。', 2)
            continue

        # チェック処理ここまで

        # caseDailycsvの左上の欄から氏名を取得
        namedf = csvConvert(val1)
        name = namedf.columns.values[0]

        # masterdf変数にmastercsvファイルから読み込んだマスターデータを格納
        # （mastercsvファイルが存在しなければ作成）
        masterdf = checkMasterCSV(valB)

        # 氏名をキーとしてマスターデータ内を検索
        keyExistFlag, id, assessmentAuthor = findMaster(masterdf, name)

        window['-name-'].update(name)

        # 存在すれば入力欄に反映
        if keyExistFlag :
            window['-id-'].update(str(id))
            window['-assessmentAuthor-'].update(str(assessmentAuthor))

            # 変更前データを変数に格納　変更チェックに使う
            dataBuffer = [name, id, assessmentAuthor]
            data_mode = 2
            statusUpdate(f'氏名：{name}の情報を取得しました。', 1)
        else :
            statusUpdate(f'氏名：{name}が登録されていません。\
受給者証番号・月間アセスメント担当者名を入力してください。', 1)
            data_mode = 1

    # クリアボタンが押された時の処理
    if event == '-clearBtn-' :
        flag = True

        # 処理時点の日付の文字列を取得　様式：○○年○○月○○日
        strToday = setToday()
        
        # テキスト入力欄を一部除き全てクリア
        # 月間アセスメント実施日の欄のみ日付を表示
        # チェックボックスはオフにする
        window['-caseDailyPath-'].update('')
        window['-userCaseDailyPath-'].update('')
        window['-assessmentDay-'].update(strToday)
        window['-outputPath-'].update('')
        window['-name-'].update('')
        window['-id-'].update('')
        window['-assessmentAuthor-'].update('')
        window['-popFillPositionCheck-'].update(value=False)
        
        # ステータス表示テキストをクリア
        statusUpdate('', 1)
        statusUpdate('', 2)

    # 実行ボタンが押された時の処理
    if event == '-submitBtn-' :
        flag = True
        val3 = values['-assessmentDay-']
        val4 = values['-outputPath-']
        val5 = values['-name-']
        val6 = values['-id-']
        val7 = values['-assessmentAuthor-']
        valA = values['-formatPath-']
        popCheckStatus = values['-popFillPositionCheck-']

        # テキストボックスおよびファイル検査

        verify_flag = True

        # buffer_flagがTrueのとき→既にファイルが存在し、上書き処理をする
        if buffer_flag :
            if buffer != [val1, val2, val3, val4, val5, val6, val7, valA, valB,
                          popCheckStatus] :buffer_flag = False
            else :
                try:
                    wb = openpyxl.load_workbook(savePath)
                except PermissionError as e:
                    verify_flag = False
                    statusUpdate('出力先のExcelブックを開くことが出来ませんでした。\n\
Excelブックを閉じてからやり直してください。', 2)
                    continue
                # マスターデータの処理はsubmit()関数内で行っている
                submit()
                buffer_flag = False
                buffer = []
                continue

        #??. 各テキスト欄が空欄でない
        if val3 == '' or val4 == '' or val5 == '' or val6 == '' or val7 == '' :
            verify_flag = False
            empty_list = []
            if val3 == '' : empty_list.append('月間アセスメント実施日')
            if val4 == '' : empty_list.append('Excelブックの出力先')
            if val5 == '' : empty_list.append('氏名')
            if val6 == '' : empty_list.append('受給者証番号')
            if val7 == '' : empty_list.append('月間アセスメント担当者名')

            empty_str = ''
            for i, j in enumerate(empty_list) :
                if i > 0 : 
                    if i == len(empty_list) - 1 :empty_str += 'および' 
                    else : empty_str += '、'
                empty_str += j
            statusUpdate(f'{empty_str}は必須項目です。空欄にはできません。', 2)

        #??. data_modeが0のとき＝データは全て埋めているが「抽出」ボタンを押してないとき
        if data_mode == 0 :
            verify_flag = False
            statusUpdate('先に「抽出」ボタンを押してください。', 2)
            continue

        if not verify_flag :continue

        #??. 出力先が絶対パスである
        if not os.path.isabs(val4) :
            verify_flag = False
            statusUpdate('Excelブックの出力先はファイル選択ダイアログから指定してください。', 2)
            continue

        #??. Formatファイルが存在しない場合エラーを出す
        #??. Formatファイルのアクセス拒否エラーをTryで渡す
        if os.path.exists(valA) :
            try:
                wb = openpyxl.load_workbook(valA)
            except PermissionError as e:
                verify_flag = False
                statusUpdate('フォーマットとなるExcelブックを開くことが出来ませんでした。\n\
Excelブックを閉じてからやり直してください。', 2)
        else :
            verify_flag = False
            statusUpdate('フォーマットとなるExcelブックが存在しません。', 2)
            
        if not verify_flag :continue
        
        #??. 出力されるxlsxファイルが既に存在する時、上書きするかどうかのメッセージを出す
        #??. 出力されるxlsxファイルのアクセス拒否エラーをTryで渡す
        #   不要. caseDailyおよびuserCaseDailyの各ファイルについてもアクセス拒否エラーをTryで渡す
        namedf = csvConvert(val1)
        name = namedf.columns.values[0]
        
        separate_code = namedf[name][0][4]
        yyyymm_list = namedf[name][0].split(separate_code)
        
        # ファイル名は'～～_○○年○○月_○○.xlsx'
        saveFilename = f'在宅支援記録表及びアセスメント表（週間・月間）_\
{yyyymm_list[0][-2:]}年{str(int(yyyymm_list[1]))}月_{name}.xlsx'
        savePath = f'{val4}/{saveFilename}'

        if os.path.exists(savePath) :
            # 上書き確認する、または、マスターデータから参照したデータが変更された時
            if not buffer_flag or (
                data_mode == 2 and data_flag == False and
                dataBuffer != [val5, val6, val7]
            ):
                errMsg_list = []
                if data_mode == 2 and data_flag == False:
                    if dataBuffer[0] != val5 :
                        errMsg_list.append(f'氏名変更：{dataBuffer[0]}→{val5}\n')
                    if dataBuffer[1] != val6 :
                        errMsg_list.append(f'受給者証番号変更：{dataBuffer[1]}→{val6}\n')
                    if dataBuffer[2] != val7 :
                        errMsg_list.append(f'月間アセスメント担当者名変更：{dataBuffer[2]}→{val7}\n')
                    errMsg_list.append('変更を適用する場合は、もう一度「実行」ボタンを押してください。\n')
                    verify_flag = False
                    data_flag = True
                    
                if not buffer_flag :
                    errMsg_list.append(f'{savePath}は既に存在します。\n\
上書きする場合は、もう一度「実行」ボタンを押してください。\n\
※上記のテキストボックスの値を変更してから「実行」ボタンを押した場合、もう一度判定を行います。')
                    verify_flag = False
                    buffer_flag = True
                    buffer = [val1, val2, val3, val4, val5, val6, val7, valA, valB, popCheckStatus]
                statusUpdate(errMsg_list, 2)

        if not verify_flag :continue

        # マスターデータの処理はsubmit()関数内で行っている
        submit()

if not flag : sys.exit(1)

